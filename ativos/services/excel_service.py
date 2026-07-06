from io import BytesIO

from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .inventario_metrics import computadores_com_alerta, contar_status, motivos_pendencia, pendencias_q
from . import configuracao_service


HEADER_FILL = PatternFill('solid', fgColor='0F2742')
HEADER_FONT = Font(color='FFFFFF', bold=True)
TITLE_FILL = PatternFill('solid', fgColor='E0F2FE')
TITLE_FONT = Font(color='0F2742', bold=True, size=14)
SUBTITLE_FONT = Font(color='64748B', italic=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='DBE3EE'),
    right=Side(style='thin', color='DBE3EE'),
    top=Side(style='thin', color='DBE3EE'),
    bottom=Side(style='thin', color='DBE3EE'),
)
STATUS_FILLS = {
    'Ativo': PatternFill('solid', fgColor='DCFCE7'),
    'Manutencao': PatternFill('solid', fgColor='FEF3C7'),
    'Manutenção': PatternFill('solid', fgColor='FEF3C7'),
    'Desligado': PatternFill('solid', fgColor='FFE4E6'),
    'Reserva': PatternFill('solid', fgColor='E0F2FE'),
}


def _local_datetime(value):
    if not value:
        return ''

    return timezone.localtime(value).strftime('%d/%m/%Y %H:%M')


def _generated_at():
    return timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')


def _status_key(status):
    return str(status or '').replace('ç', 'c').replace('ã', 'a').replace('Ã§', 'c')


def _append_title(ws, title, subtitle=None, columns=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    ws['A1'] = title
    ws['A1'].fill = TITLE_FILL
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
        ws['A2'] = subtitle
        ws['A2'].font = SUBTITLE_FONT
        ws['A2'].alignment = Alignment(horizontal='center')


def _style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def _style_table(ws, header_row=1):
    _style_header(ws[header_row])
    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            if cell.column_letter in {'A', 'B', 'C'}:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

            if ws.cell(row=header_row, column=cell.column).value == 'Status':
                fill = STATUS_FILLS.get(cell.value) or STATUS_FILLS.get(_status_key(cell.value))
                if fill:
                    cell.fill = fill


def _auto_width(ws):
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 60))

        ws.column_dimensions[letter].width = max(12, max_length + 3)


def _finish_sheet(ws, header_row=1):
    _style_table(ws, header_row=header_row)
    _auto_width(ws)


def _add_resumo(wb, computadores):
    ws = wb.active
    ws.title = 'Resumo'
    alertas = computadores_com_alerta(computadores)
    status = contar_status(computadores)
    rows = [
        ('Total de computadores', computadores.count()),
        ('Ativos', status['ativos']),
        ('Manutencao', status['manutencao']),
        ('Desligados', status['desligados']),
        ('Reservas', status['reservas']),
        ('Total de setores', computadores.exclude(sala='').values('sala').distinct().count()),
        ('Computadores com alerta', len(alertas)),
        ('Data de geracao', _generated_at()),
    ]

    empresa = configuracao_service.get_config().nome_empresa or 'JR Grupo'
    _append_title(ws, 'Gestao de Ativos de TI', f'{empresa} - Resumo geral do inventario', columns=2)
    ws.append([])
    ws.append(['Indicador', 'Valor'])
    for row in rows:
        ws.append(row)

    _finish_sheet(ws, header_row=4)


def _add_computadores(wb, computadores):
    ws = wb.create_sheet('Computadores')
    ws.append(
        [
            'ID',
            'Sala',
            'Status',
            'Usuario',
            'Sistema',
            'RAM',
            'Processador',
            'Armazenamento',
            'Placa de video',
            'Observacoes',
            'Atualizado em',
        ]
    )
    for computador in computadores.order_by('sala', 'id'):
        ws.append(
            [
                computador.id,
                computador.sala,
                computador.status,
                computador.usuario,
                computador.sistema,
                computador.ram,
                computador.processador,
                computador.armazenamento,
                computador.placa_video,
                computador.observacoes,
                _local_datetime(computador.atualizado_em),
            ]
        )

    _finish_sheet(ws)


def _add_pendencias(wb, computadores):
    ws = wb.create_sheet('Pendencias')
    ws.append(['ID', 'Sala', 'Status', 'Usuario', 'Pendencias'])

    for computador in computadores.filter(pendencias_q()).distinct().order_by('sala', 'id'):
        ws.append(
            [
                computador.id,
                computador.sala,
                computador.status,
                computador.usuario,
                ', '.join(motivos_pendencia(computador)),
            ]
        )

    _finish_sheet(ws)


def _add_alertas(wb, computadores):
    ws = wb.create_sheet('Alertas')
    ws.append(['ID', 'Sala', 'Status', 'Usuario', 'RAM', 'Motivos'])

    for item in computadores_com_alerta(computadores):
        computador = item['computador']
        ws.append(
            [
                computador.id,
                computador.sala,
                computador.status,
                computador.usuario,
                computador.ram,
                ', '.join(item['motivos']),
            ]
        )

    _finish_sheet(ws)


def _add_resumo_setor(wb, computadores):
    ws = wb.create_sheet('Resumo por setor')
    ws.append(['Sala', 'Total', 'Ativos', 'Manutencao', 'Desligados', 'Reservas', 'Alertas'])

    salas = computadores.exclude(sala='').values_list('sala', flat=True).distinct().order_by('sala')
    for sala in salas:
        qs = computadores.filter(sala=sala)
        status = contar_status(qs)
        ws.append(
            [
                sala,
                qs.count(),
                status['ativos'],
                status['manutencao'],
                status['desligados'],
                status['reservas'],
                len(computadores_com_alerta(qs)),
            ]
        )

    _finish_sheet(ws)


def _add_movimentacoes(wb, movimentacoes):
    ws = wb.create_sheet('Movimentacoes')
    ws.append(['Data/hora', 'Computador', 'Acao', 'Campo', 'Valor anterior', 'Valor novo', 'Usuario responsavel'])

    for mov in movimentacoes:
        ws.append(
            [
                _local_datetime(mov.data_hora),
                mov.computador_id or mov.computador_identificador,
                mov.acao,
                mov.campo,
                mov.valor_anterior,
                mov.valor_novo,
                mov.usuario_responsavel,
            ]
        )

    _finish_sheet(ws)


def inventario_completo_xlsx(computadores, movimentacoes):
    empresa = configuracao_service.get_config().nome_empresa or 'JR Grupo'
    wb = Workbook()
    wb.properties.title = 'Inventario de Ativos de TI'
    wb.properties.creator = empresa

    _add_resumo(wb, computadores)
    _add_computadores(wb, computadores)
    _add_pendencias(wb, computadores)
    _add_alertas(wb, computadores)
    _add_resumo_setor(wb, computadores)
    _add_movimentacoes(wb, movimentacoes)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
